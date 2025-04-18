import os.path
import openpyxl
import json
import zipfile
from pathlib import Path
from offlinesec_client.const import SUBDIR

import offlinesec_client.func

FILENAME = "role_masking.json"


class Agr1251:
    def __init__(self, filename):
        self.filename = filename
        self.data = list()

    def read_xlsx_file(self):
        self.data = list()
        wb_obj = openpyxl.load_workbook(self.filename)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column

        max_row = sheet_obj.max_row
        for i in range(2, max_row + 1):
            deleted = sheet_obj.cell(row=i, column=11).value
            if deleted.strip() == "X":
                continue

            new_item = dict()
            new_item["mandt"] = sheet_obj.cell(row=i, column=1).value
            new_item["agr_name"] = sheet_obj.cell(row=i, column=2).value
            new_item["object"] = sheet_obj.cell(row=i, column=4).value
            new_item["auth"] = sheet_obj.cell(row=i, column=5).value
            new_item["field"] = sheet_obj.cell(row=i, column=7).value
            new_item["low"] = sheet_obj.cell(row=i, column=8).value
            new_item["high"] = sheet_obj.cell(row=i, column=9).value

            self.data.append(new_item)

    def read_txt_file(self):
        self.data = list()
        f = open(self.filename, "r")

        header_flag = True
        for line in f:
            if line.startswith("|"):
                if header_flag:
                    header_flag = False
                else:
                    splited_line = line.split("|")
                    deleted = splited_line[11].strip()
                    if deleted.strip() == "X":
                        continue
                    new_item = dict()
                    new_item["mandt"] = splited_line[1].strip()
                    new_item["agr_name"] = splited_line[2].strip()
                    new_item["object"] = splited_line[4].strip()
                    new_item["auth"] = splited_line[5].strip()
                    new_item["field"] = splited_line[7].strip()
                    new_item["low"] = splited_line[8].strip()
                    new_item["high"] = splited_line[9].strip()

                    self.data.append(new_item)

        f.close()

    def read_file(self):
        if self.filename.split(".")[-1].lower() == "xlsx":
            self.read_xlsx_file()
        elif self.filename.split(".")[-1].lower() == "txt":
            self.read_txt_file()

    def masking(self):
        ROLE_MASK_TEMPLATE = "Z_ROLE_%s"
        mask_base = Agr1251.read_mask()

        for line in self.data:
            if line["agr_name"] in mask_base:
                idx = mask_base.index(line["agr_name"])
                line["agr_name"] = ROLE_MASK_TEMPLATE % (str(idx))
            else:
                pos = len(mask_base)
                mask_base.append(line["agr_name"])
                line["agr_name"] = ROLE_MASK_TEMPLATE % (str(pos))

        Agr1251.save_mask(mask_base)

    def save_results(self):
        filename = offlinesec_client.func.get_file_name("roles.json")
        with open(filename, "w") as f:
            json.dump(self.data, f)

        zipfilename = offlinesec_client.func.get_file_name("roles.zip")
        zipfile.ZipFile(zipfilename, mode='w', compression=zipfile.ZIP_DEFLATED, compresslevel=9).write(filename)

        if os.path.isfile(filename):
            os.remove(filename)
        return zipfilename

    @staticmethod
    def save_mask(mask_base):
        folder = os.path.join(Path.home(), SUBDIR)
        os.makedirs(folder, exist_ok=True)
        filename = os.path.join(folder, FILENAME)
        with open(filename, "w") as f:
            json.dump(mask_base, f)

    @staticmethod
    def read_mask():
        filename = os.path.join(Path.home(), SUBDIR, FILENAME)

        if not os.path.isfile(filename):
            return list()

        with open(filename, "r") as f:
            return json.load(f)
