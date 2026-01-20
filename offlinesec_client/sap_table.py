# -*- coding: utf-8 -*-

import os.path
import openpyxl
from offlinesec_client.sap_table_def import COLUMN_REPLACEMENT, DATA_REPLACEMENT

LONG_SAP_TABLES = ["ICFSERVICE"]

class SAP_Table_SE16:
    def __init__(self, table_name, file_name):
        self.table_name = table_name
        self.file_name = file_name

        self.columns = list()
        self.data = list()

        self.read_table()

        if table_name:
            if table_name in COLUMN_REPLACEMENT.keys():
                self.replace_columns(COLUMN_REPLACEMENT[table_name])

    def replace_columns(self, replacement):
        for column in replacement:
            for item in replacement[column]:
                if item in self.columns:
                    idx = self.columns.index(item)
                    self.columns[idx] = column
                    #continue
                    break

    def replace_data_values(self, new_line):
        if self.table_name in DATA_REPLACEMENT:
            for column in DATA_REPLACEMENT[self.table_name]:
                column_idx = self.columns.index(column)
                value = new_line[column_idx]
                for new_value in DATA_REPLACEMENT[self.table_name][column]:
                    if value in DATA_REPLACEMENT[self.table_name][column][new_value]:
                        new_line[column_idx] = new_value
                        break

        return new_line

    def read_table(self):
        if not self.file_name:
            raise ValueError("File name is empty")

        if not os.path.isfile(self.file_name):
            raise ValueError("File %s not exist" % (self.file_name,))

        if not self.file_name.lower().endswith(".txt"):
            raise ValueError("Bad file extension %s" % (self.file_name,))

        f = open(self.file_name, "r", errors="ignore")
        header_flag = True
        column_num = 0
        column_lens = list()

        for line in f:
            if line.startswith("|"):
                if header_flag:
                    header_flag = False
                    self.columns = [item.strip() for item in line.split("|")]
                    column_lens = [len(item) for item in line.split("|")]
                    if not len(self.columns):
                        raise ValueError("Wrong file structure. Columns not found in the file")
                    column_num = len(self.columns)
                else:
                    row = [item.strip() for item in line.split("|")]
                    if len(row) != column_num:
                        row = list()
                        pos = 0
                        for i in range(0, len(self.columns)):
                            column_len = column_lens[i]
                            row.append(line[pos:pos+column_len].strip())
                            pos += column_len + 1

                        if len(row) != column_num:
                            raise ValueError("Wrong file structure. Bad line: %s in the file" % (line,))

                    self.data.append(row)

        f.close()

        if not len(self.columns):
            raise ValueError("Wrong file structure. Columns not found in the file")

    def __iter__(self):
        for row in self.data:
            out_row = dict()
            for num, column in enumerate(self.columns):
                out_row[column] = row[num]
            yield out_row


class SAP_Long_Table_SE16(SAP_Table_SE16):
    def read_table(self):
        if not self.file_name:
            raise ValueError("File name is empty")

        if not os.path.isfile(self.file_name):
            raise ValueError("File %s not exist" % (self.file_name,))

        if not self.file_name.lower().endswith(".txt"):
            raise ValueError("Bad file extension %s" % (self.file_name,))

        f = open(self.file_name, "r", errors="ignore")
        column_num = 0
        column_lens = list()
        self.columns = list()
        separator = 0
        buffer = list()
        lines_per_raw = 0

        for line in f:
            if len(line.strip()) > 0 and line.strip() == "-" * len(line.strip()):
                separator += 1

                if separator == 1 :
                    buffer = list()

                elif separator == 2:
                    column_line = "".join(buffer)
                    self.columns.extend([item.strip() for item in column_line.split("|")])
                    column_lens = [len(item) for item in column_line.split("|")]
                    if not len(self.columns):
                        raise ValueError("Wrong file structure. Columns not found in the file")
                    column_num = len(self.columns)
                    buffer = list()
                    #print("columns: %s, length: %s" % (len(self.columns), len(column_line)))
                    keys = dict()
                    for i in range(0, len(self.columns)):
                        keys[i] = column_lens[i]
                    #print(keys)
                continue
            elif 1 <= separator < 2:
                buffer.append(line)
                #print(line, len(line))
                lines_per_raw += 1

            elif separator >= 2:
                if line.strip() == "":
                    continue
                buffer.append(line)
                #print(len(line))

                if len(buffer) >= lines_per_raw:
                    new_line = "".join(buffer)
                    row = [item for item in new_line.split("|")]

                    if len(row) > column_num:
                        row = list()
                        pos = 0

                        # "|" inside
                        for i in range(0, len(self.columns)):
                            column_len = column_lens[i]
                            row.append(new_line[pos:pos + column_len].strip())

                            pos += column_len + 1

                        new_item = dict()
                        for i in range(0, len(self.columns)):
                            new_item[self.columns[i]] = row[i]
                        #print(new_item)

                        if len(row) != column_num:
                            raise ValueError("Wrong file structure. Bad line: %s in the file" % (new_line,))

                    self.data.append(row)
                    buffer = list()

        f.close()

        if not len(self.columns):
            raise ValueError("Wrong file structure. Columns not found in the file")


class HANA_SAP_Table(SAP_Table_SE16):
    def read_table(self):
        if not self.file_name:
            raise ValueError("File name is empty")

        if not os.path.isfile(self.file_name):
            raise ValueError("File %s not exist" % (self.file_name,))

        if not self.file_name.lower().endswith(".txt"):
            raise ValueError("Bad file extension %s" % (self.file_name,))

        f = open(self.file_name, "r", errors="ignore")
        header_flag = True
        column_num = 0

        for line in f:
            if True:
                if header_flag:
                    header_flag = False
                    self.columns = [item.strip() for item in line.split(";")]
                    if len(self.columns) < 2:
                        raise ValueError("Wrong file structure. Columns not found in the file")
                    column_num = len(self.columns)
                else:
                    row = [item.strip() for item in line.split(";")]
                    if len(row) != column_num:
                        raise ValueError("Wrong file structure. Bad line: %s in the file" % (line,))
                    self.data.append(row)

        f.close()


class XLSX_SE16_Table(SAP_Table_SE16):
    def read_table(self):
        if self.file_name is None or self.file_name == "":
            raise ValueError("File name is empty")

        if not os.path.isfile(self.file_name):
            raise ValueError("File %s not exist" % (self.file_name,))

        if not self.file_name.lower().endswith(".xlsx"):
            raise ValueError("Bad file extension %s" % (self.file_name,))

        wb_obj = openpyxl.load_workbook(self.file_name)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        column_num = 0

        for i in range(1, max_row + 1):
            if i == 1:
                for k in range(1, max_col + 1):
                    self.columns.append(sheet_obj.cell(row=i, column=k).value)
                if not len(self.columns):
                    raise ValueError("Wrong file structure. Columns not found")
                else:
                    column_num = len(self.columns)
                    if self.table_name in COLUMN_REPLACEMENT.keys():
                        self.replace_columns(COLUMN_REPLACEMENT[self.table_name])
            else:
                new_line = list()
                for k in range(1, max_col + 1):
                    new_line.append(sheet_obj.cell(row=i, column=k).value)
                if len(new_line) != column_num:
                    raise ValueError("Wrong file structure. Bad line: %s in the file" % (", ".join(new_line,)))

                self.data.append(self.replace_data_values(new_line))

        wb_obj.close()


class SAPTable:
    def __init__(self, table_name, file_name, sys_name=None):
        classmap = {
            "SAP_Table_SE16": SAP_Table_SE16,
            "HANA_SAP_Table": HANA_SAP_Table,
            "XLSX_SE16_Table": XLSX_SE16_Table,
        }

        self.table_name = table_name
        self.file_name = file_name
        self.sys_name = sys_name

        created_object = None
        self.err_list = list()
        for class_name in classmap:
            try:
                if class_name == "SAP_Table_SE16" and table_name in LONG_SAP_TABLES:
                    created_object = SAP_Long_Table_SE16(table_name, file_name)
                    self.class_name = class_name
                else:
                    created_object = classmap[class_name](table_name, file_name)
                    self.class_name = class_name
            except ValueError as err:
                self.err_list.append(class_name + ": " + str(err))
            else:
                break

        if created_object:
            self.tbl = created_object
            self.columns = self.tbl.columns
            self.data = self.tbl.data

        else:
            raise ValueError("Unsupported file format")


    def __iter__(self):
        for row in self.tbl.data:
            out_row = dict()
            for num, column in enumerate(self.tbl.columns):
                out_row[column] = row[num]
            yield out_row

    def check_columns(self):
        if self.table_name in COLUMN_REPLACEMENT:
            for column in COLUMN_REPLACEMENT[self.table_name]:
                if column not in self.columns:
                    if self.sys_name:
                        self.err_list.append("The required column %s not found in the file '%s' (the %s table in %s)" % (
                            column, self.file_name, self.table_name, self.sys_name))
                    else:
                        self.err_list.append("The required column %s not found in the file '%s' (the %s table)" % (
                            column, self.file_name, self.table_name))
                    return False

        return True
