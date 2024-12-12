import editpyxl
import openpyxl
import os
import argparse
from pathlib import Path

from offlinesec_client.const import SUBDIR, FILE
from offlinesec_client.masking import Masking, SAPSID_MASK, USER_MASK,HOST_MASK, PATH_MASK, RFCDEST_MASK
import offlinesec_client.func
import json

ROLE_MASK_TEMPLATE = "Z_ROLE_%s"
FILENAME = "role_masking.json"



def main():
    args = init_args()
    if FILE in args and args[FILE]:
        file = args[FILE]
        read_file(file)


def check_file_arg(s):
    return offlinesec_client.func.check_file_arg(s, ["xlsx"], 0)


def init_args():
    parser = argparse.ArgumentParser()

    parser.add_argument("-f", "--%s" % (FILE,), action="store", type=check_file_arg,
                        help="Report file", required=True)

    parser.parse_args()
    return vars(parser.parse_args())

def do_rfc_transform(file_name, sheet_name="Parameters"):
    wb = editpyxl.Workbook()
    source_filename = file_name

    wb.open(source_filename)

    if not sheet_name in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]

    masking = {
        SAPSID_MASK: Masking(SAPSID_MASK),
        USER_MASK: Masking(USER_MASK),
        HOST_MASK: Masking(HOST_MASK),
        PATH_MASK: Masking(PATH_MASK),
        RFCDEST_MASK: Masking(RFCDEST_MASK)
    }

    column_map = {
        "Source SID": SAPSID_MASK,
        "RFC Destination": RFCDEST_MASK,
        "Destination SID": SAPSID_MASK,
        "Source Host": HOST_MASK,
        "User in Destination System": USER_MASK,
        "Proxy": HOST_MASK,
        "URL": PATH_MASK
    }


    max_masking = 20
    columns = list()
    for i in range(1, max_masking + 1):
        try:
            column_name = ws.cell(row=1, column=i).value
            if column_name is None or column_name.strip() == "":
                break
            if column_name in column_map:
                columns.append(column_map[column_name])
            else:
                columns.append(None)
                continue
        except IndexError as err:
            break

    max_lines = 10000
    for i, column in enumerate(columns):
        for k in range(2, max_lines + 1):
            if column is not None:
                try:
                    cell = ws.cell(row=k, column=i + 1).value

                except IndexError as err:
                    break
                else:
                    if cell is not None and cell.strip() != "":
                        new_value = masking[column].do_unmask(cell)
                        ws.cell(row=k, column=i + 1).value = new_value
                    else:
                        break

    wb.save(fullFilename=file_name)
    wb.close()



def do_secnotes_transform(file_name):
    wb = editpyxl.Workbook()
    source_filename = file_name

    wb.open(source_filename)

    if not "Parameters" in wb.sheetnames:
        wb.close()
        return

    ws = wb["Parameters"]
    sapsid_masking = Masking(SAPSID_MASK)
    max_lines = 10000
    for i in range(2, max_lines):
        try:
            cell = ws.cell(row=i, column=1).value
        except IndexError as err:
            break
        else:
            if cell is not None and cell.strip() != "":
                new_value = sapsid_masking.do_unmask(cell)
                ws.cell(row=i, column=1).value = new_value
            else:
                break

    wb.save(fullFilename=file_name)
    wb.close()


def read_file(file):
    flag = False
    if os.path.basename(file).startswith("roles_"):
        flag = True
    elif os.path.basename(file).startswith("rfc_") and os.path.basename(file).endswith(".xlsx"):
        do_rfc_transform(file)
    elif os.path.basename(file).startswith("secnotes_") and os.path.basename(file).endswith(".xlsx"):
        do_secnotes_transform(file)

    if not flag:
        return
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    title = ws['B1'].value

    if title.startswith("Critical Authorizations"):
        roles = read_dict_file()
        resolve_roles_page1(ws, roles)
        ws = wb.worksheets[1]
        resolve_roles_page2(ws, roles)


        wb.save(file)
        print(" * Roles in file %s have been converted" % (os.path.basename(file),))



def resolve_roles_page1(ws, roles):
    max_row = ws.max_row
    for i in range(5, max_row + 1):
        cell = ws.cell(row=i, column=1).value
        if cell and cell.startswith(ROLE_MASK_TEMPLATE[:6]):
            try:
                ind = int(cell[7:])
                if ind <= len(roles):
                    ws.cell(row=i, column=1).value = roles[ind]
            except:
                continue


def resolve_roles_page2(ws,roles):
    max_row = ws.max_row
    for i in range(2, max_row + 1):
        cell = ws.cell(row=i, column=2).value
        if cell and cell.startswith(ROLE_MASK_TEMPLATE[:6]):
            try:
                ind = int(cell[7:])
                if ind <= len(roles):
                    ws.cell(row=i, column=2).value = roles[ind]
            except:
                continue


def read_dict_file():
    filename = os.path.join(Path.home(), SUBDIR, FILENAME)

    if not os.path.isfile(filename):
        return list()

    with open(filename, "r") as f:
        return json.load(f)


if __name__ == '__main__':
    main()
