import json
import os.path
import re
import openpyxl
import zipfile
import offlinesec_client.func
from offlinesec_client.const import PARAM_NAME, PARAM_DESC, PARAM_VALUE

HIDE_PARAMS = ["FN_BDCALTLOG", "FN_BDCLOG", "FN_EXTRACT", "SAPARGV", "SAPGLOBALHOST", "SAPLOCALHOST",
               "SAPLOCALHOSTFULL", "SAPPROFILE", "SAPPROFILE_IN_EFFECT", "SAPSYSTEMNAME", "SETENV_.*", "Execute_.*",
               "Start_Program_.*", r"_\S{2}", "dbs/hdb/schema", "dbs/mss/dbname", "dbs/ora/tnsname", "dbs/syb/dbname",
               "enq/server/schema_0", "enque/encni/hostname", "enque/serverhost", "igs/listener/rfc", "igs/mux/ip",
               "ms/comment", "rdisp/j2ee_profile", "rdisp/j2ee_profile", "rdisp/mshost", "rdisp/msserv", "rdisp/myname",
               "rlfw/bri/msserv", "rlfw/upg/msserv", "snc/gssapi_lib", "snc/identity/as", "vmcj/debug_proxy/cfg/msHost",
               "vmcj/debug_proxy/cfg/msPort"
               ]

COLUMNS_ENG = [
    "Parameter Name",
    "User-Defined Value",
    "System Default Value",
    "System Default Value(Unsubstituted Form)",
    "Comment"]


class ProfileParamList:
    def __init__(self):
        self.params = list()

    @staticmethod
    def param_name(name):
        return name.strip()

    def add(self, name, user_defined_value, system_defined="", third_value="", desc=""):
        if not name:
            return

        new_item = dict()
        new_item[PARAM_NAME] = ProfileParamList.param_name(name)
        new_item[PARAM_VALUE] = ProfileParamList.param_value(user_defined_value, system_defined, third_value)
        if desc:
            new_item[PARAM_DESC] = desc

        self.params.append(new_item)


class RsparamReport:
    def __init__(self, filename):
        self.filename = filename
        self.format = False
        self.data = dict()
        self.read_file()

    @staticmethod
    def mask(param_name, param_value):
        DEFAULT_VALUE = "XXX"
        for hide_param in HIDE_PARAMS:
            re_str = "(?i)^" + hide_param.lower() + "$"
            res = re.match(re_str, param_name)
            if res:
                return DEFAULT_VALUE if param_value != "" else ""

        param_value = re.sub(r'\/usr\/sap\/\S{3}\/', '/usr/sap/XXX/', param_value)
        return param_value

    def read_file(self):
        #try
        wb_obj = openpyxl.load_workbook(self.filename)

        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column

        if max_col == 5:
            temp_list = list()
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=1, column=i)
                temp_list.append(cell_obj.value)
            if len(temp_list) == 5:
                self.format = self.check_format(temp_list)
        else:
            raise AttributeError("Wrong file format. The file contains only %s columns" % (max_col,))

        if self.format:
            self.parse_data(sheet_obj)
        else:
            raise AttributeError("Wrong file format. Check column names in the file")

        wb_obj.close()

    @staticmethod
    def check_format(columns):
        for i in range(0, 5):
            if columns[i] != COLUMNS_ENG[i]:
                return False
        return True

    def parse_data(self, sheet_obj,):
        max_row = sheet_obj.max_row
        for i in range(2, max_row + 1):
            col1 = sheet_obj.cell(row=i, column=1).value
            col2 = sheet_obj.cell(row=i, column=2).value
            col3 = sheet_obj.cell(row=i, column=3).value
            col4 = sheet_obj.cell(row=i, column=4).value
            col5 = sheet_obj.cell(row=i, column=5).value

            col1 = col1 if col1 is not None else ""
            col2 = col2 if col2 is not None else ""
            col3 = col3 if col3 is not None else ""
            col4 = col4 if col4 is not None else ""
            col5 = col5 if col5 is not None else ""

            self.add_param(col1, col2, col3, col4, col5)

    def add_param(self, param_name, user_value, system_value, third_value, descr=""):
        name = RsparamReport.param_name(param_name)
        value = RsparamReport.param_value(user_value, system_value, third_value)
        value = RsparamReport.mask(name, value)
        if name not in self.data.keys():
            self.data[name] = (value, descr)

    @staticmethod
    def param_name(name):
        return name.strip().lower()

    @staticmethod
    def param_value(user_defined_value, system_defined="", third_value=""):
        if user_defined_value is not None and user_defined_value.strip() != "":
            return user_defined_value.strip()
        elif system_defined is not None and system_defined.strip() != "":
            return system_defined.strip()
        elif third_value is not None and third_value.strip() != "":
            return third_value.strip()
        else:
            return ""

    def save_to_file(self):
        filename = offlinesec_client.func.get_file_name("params.json")

        with open(filename, 'w') as f:
            json.dump(self.data, f)

        zipfilename = offlinesec_client.func.get_file_name("params.zip")
        zipfile.ZipFile(zipfilename, mode='w', compression=zipfile.ZIP_DEFLATED, compresslevel=9).write(filename)

        if os.path.isfile(filename):
            os.remove(filename)

        return zipfilename
