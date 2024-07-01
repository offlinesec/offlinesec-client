import openpyxl
import re


class Cwbntcust:
    def __init__(self, filename):
        self.filename = filename
        self.data = list()

    def read_file(self):
        if self.filename.split(".")[-1].lower() == "xlsx":
            return self.read_xlsx_file()
        elif self.filename.split(".")[-1].lower() == "txt":
            return self.read_txt_file()

    def read_xlsx_file(self):
        self.data = list()
        outlist = list()
        wb_obj = openpyxl.load_workbook(self.filename)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column

        max_row = sheet_obj.max_row
        for i in range(2, max_row + 1):
            note = str(sheet_obj.cell(row=i, column=1).value)
            ntstatus = str(sheet_obj.cell(row=i, column=2).value)
            prstatus = str(sheet_obj.cell(row=i, column=3).value)
            if Cwbntcust.check_note_status(ntstatus=ntstatus, prstatus=prstatus):
                outlist.append(note)
        return outlist

    @staticmethod
    def check_note_status(ntstatus, prstatus):
        if prstatus == "E" or prstatus == "O":              # Completely implemented or obsolete
            return True
        return False

    def read_txt_file(self):
        outlist = list()
        f = open(self.filename, "r")
        header_flag = True
        note_column = 0
        for line in f:
            if line.startswith("|"):
                if header_flag:
                    header_flag = False
                else:
                    splited_line = line.split("|")
                    if note_column == 0:
                        for i in range(0, 2):
                            result = re.match(r"\d{10}", str(splited_line[i]).strip())
                            if result:
                                note_column = i
                                break
                    if note_column:
                        note = str(splited_line[note_column].strip())
                        ntstatus = str(splited_line[note_column + 1].strip())
                        prstatus = str(splited_line[note_column + 2].strip())
                        if Cwbntcust.check_note_status(ntstatus=ntstatus, prstatus=prstatus):
                            outlist.append(note)

        f.close()
        return outlist
